import streamlit as st
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
from urllib3.exceptions import InsecureRequestWarning
from io import BytesIO
import tempfile



output = BytesIO()

KEY = st.text_input('Enter CDX API Key', '')
st.write("API key is " + KEY)

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

uploaded_file = st.file_uploader("Upload .xlsx file here", type="xlsx", accept_multiple_files=False, key=None, help=None, on_change=None, args=None, kwargs=None, disabled=False)

def find_zip_col(worksheet):
    for row_cells in worksheet.iter_rows():
            for cell in row_cells:
                #print(cell.value)
                if cell: cell_str = str(cell.value)
                if cell_str and str(cell_str.split("-")[0]).isnumeric() and (len(cell_str) == 5 or len(cell_str) == 10):
                    return int(cell.column)
    return -1


def is_zipcode(zipcode):
    if zipcode: cell_str = str(zipcode)
    return cell_str and str(cell_str.split("-")[0]).isnumeric() and (len(cell_str) == 5 or len(cell_str) == 10)



@st.experimental_memo
def handleFile(file_path):
    workbook = load_workbook(file_path, data_only=True)
    
    for sheet in workbook.worksheets:
        start_zip_col = find_zip_col(sheet)
        row_count = sheet.max_row
        #print(start_zip_col)
        
        for row in sheet.iter_rows():
            #print()
            if row[start_zip_col -1 ].value and is_zipcode(row[start_zip_col -1 ].value):
                zipcode = int(str(row[start_zip_col -1 ].value).split("-")[0])
                URL = f"""https://geodata.cdxtech.com/api/geogeneral?key={KEY}&zipcode={zipcode}&format=json"""
                res = requests.get(URL, verify=False)
                api_city = None
                if res.json() and res.json()["results"]:
                    api_city = res.json()["results"]["city"].lower()
                if row[start_zip_col - 3 ]:
                    data_city = row[start_zip_col - 3 ].value.lower()
                output = sheet.cell(row = row[0].row, column=start_zip_col + 1)
                if api_city and api_city == data_city:
                    output.value = "True"
                elif api_city:
                    output.value = "False"
                    output = sheet.cell(row = row[0].row, column=start_zip_col + 2)
                    output.value = api_city

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        workbook.save(tmp.name)
        data = BytesIO(tmp.read())
        tmp.flush()
        tmp.close()
    return data




if uploaded_file is not None:
    
    data = handleFile(uploaded_file)
    st.download_button(
    label="Download Excel workbook",
    data=data,
    file_name="workbook.xlsx",
    mime="xlsx",
)


